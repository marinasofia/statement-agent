"""Excel workbook writer for batch results.

One writer, used by run_batch. Reconciled rows are grouped into one sheet
per statement month. Rows that did not reconcile go to a "Needs review"
sheet with the reasons and the balance delta, so a reviewer sees them
instead of losing them.

Deduplication: a statement is the same statement when it is for the same
account, the same statement date, and the same closing balance. That key
is written to a hidden "Row key" column on every sheet, so a later run
can dedup against rows already on disk no matter which columns a client
chose to display. Keying on the account holder alone would merge two
accounts owned by one person, and keying on displayed columns would make
dedup depend on layout.
"""

REVIEW_SHEET = "Needs review"
REVIEW_EXTRA_HEADERS = ["Reasons", "Balance delta", "Source file"]
ROW_KEY_HEADER = "Row key"


def _fold(value) -> str:
    return " ".join(str(value or "").casefold().split())


def _digits(value) -> str:
    return "".join(ch for ch in str(value or "") if ch.isalnum()).casefold()


def row_key(data: dict) -> str:
    account = _digits(data.get("account_number")) or _fold(data.get("account_holder"))
    closing = data.get("closing_balance")
    closing_text = f"{closing:.2f}" if isinstance(closing, (int, float)) else ""
    return f"{account}|{data.get('statement_date') or ''}|{closing_text}"

import logging
import os

from openpyxl import Workbook, load_workbook

from core.client_config import load_format_config

logger = logging.getLogger(__name__)


def _existing_keys(ws):
    header = [c.value for c in ws[1]]
    if ROW_KEY_HEADER not in header:
        return set()
    idx = header.index(ROW_KEY_HEADER)
    return {row[idx] for row in ws.iter_rows(min_row=2, values_only=True) if row[idx]}


def _hide_row_key(ws, header_count):
    letter = ws.cell(row=1, column=header_count).column_letter
    ws.column_dimensions[letter].hidden = True


def _autofit(ws, columns, label_by_name):
    for i, col_name in enumerate(columns, start=1):
        values = [label_by_name.get(col_name, col_name)]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[i - 1] is not None:
                values.append(str(row[i - 1]))
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = (
            int(max(len(v) for v in values) * 1.2) + 4
        )


def write_workbook(results, output_path, client_id, fallback_month=None):
    """Write every extracted result to the workbook.

    Returns a dict of counts: ok, needs_review, duplicate, failed.
    """
    counts = {"ok": 0, "needs_review": 0, "duplicate": 0, "failed": 0}
    workbook_dirty = False
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    if os.path.exists(output_path):
        wb = load_workbook(output_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    keys_by_sheet = {}
    columns_by_sheet = {}

    for result in results:
        if result.get("error"):
            counts["failed"] += 1
            continue

        config = load_format_config(client_id, result.get("format_id") or "default")
        columns = config["excel_output"]["columns"]
        label_by_name = {f["name"]: f.get("label", f["name"]) for f in config["fields"]}
        headers = [label_by_name.get(col, col) for col in columns]

        data = result.get("validated_data", {})
        needs_review = result.get("status") == "NEEDS_REVIEW"
        if needs_review:
            sheet = REVIEW_SHEET
            headers = headers + REVIEW_EXTRA_HEADERS
        else:
            sheet = result.get("statement_month") or fallback_month or "Unsorted"
        headers = headers + [ROW_KEY_HEADER]

        if sheet not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet)
            ws.append(headers)
            ws.freeze_panes = "A2"
            _hide_row_key(ws, len(headers))
            keys_by_sheet[sheet] = set()
        else:
            ws = wb[sheet]
            keys_by_sheet.setdefault(sheet, _existing_keys(ws))
        columns_by_sheet[sheet] = (columns, label_by_name)

        key = row_key(data)
        if key in keys_by_sheet[sheet]:
            counts["duplicate"] += 1
            result["duplicate"] = True
            logger.info("Skipping duplicate row in %s", sheet)
            continue

        row = [data.get(col) for col in columns]
        if needs_review:
            delta = (result.get("reconciliation") or {}).get("balance_delta")
            row += ["; ".join(result.get("review_reasons") or []), delta,
                    os.path.basename(result.get("file_path") or "")]
            counts["needs_review"] += 1
        else:
            counts["ok"] += 1
        row.append(key)
        ws.append(row)
        keys_by_sheet[sheet].add(key)
        workbook_dirty = True

    for sheet, (columns, label_by_name) in columns_by_sheet.items():
        extra = REVIEW_EXTRA_HEADERS if sheet == REVIEW_SHEET else []
        _autofit(wb[sheet], columns + extra, label_by_name)

    if workbook_dirty or not os.path.exists(output_path):
        wb.save(output_path)
    return counts
