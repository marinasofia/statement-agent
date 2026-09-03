import argparse
import os
import glob
import logging
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook, load_workbook
from core.config import EXCEL_OUTPUT_PATH, ALLOWED_UPLOAD_DIR
from core.client_config import get_client_id, load_format_config
from agents.statement_extraction.graph import compiled_graph

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
logger = logging.getLogger(__name__)

DEFAULT_WORKERS = 5
MONTH_PATTERN = re.compile(r"^\d{4}-(0[1-9]|1[0-2])$")


def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="Extract every PDF in a folder into one Excel workbook."
    )
    parser.add_argument("--input", default=ALLOWED_UPLOAD_DIR,
                        help="Folder of PDF statements (default: the allowed upload dir)")
    parser.add_argument("--output", default=EXCEL_OUTPUT_PATH,
                        help="Workbook to create or append to")
    parser.add_argument("--month", default=None,
                        help="Sheet to use for statements that carry no parseable date, as YYYY-MM. "
                             "Without it those rows land on an 'Unsorted' sheet.")
    parser.add_argument("--workers", type=int, default=DEFAULT_WORKERS,
                        help=f"Concurrent extractions (default {DEFAULT_WORKERS})")
    args = parser.parse_args(argv)
    if args.month and not MONTH_PATTERN.match(args.month):
        parser.error("--month must look like YYYY-MM")
    if args.workers < 1:
        parser.error("--workers must be at least 1")
    return args

def process_pdf(pdf_path: str) -> dict:
    try:
        result = compiled_graph.invoke({'file_path': pdf_path, 'skip_excel': True})
        return result
    except Exception as e:
        logger.error(f"Unhandled error on {pdf_path}: {e}")
        return {'file_path': pdf_path, 'error': str(e)}

def write_excel(results: list, fallback_month: str = None, output_path: str = EXCEL_OUTPUT_PATH):
    """Write all results to Excel in one pass."""
    client_id = get_client_id()
    config = load_format_config(client_id, "default")
    columns = config.get("excel_output", {}).get("columns", [])
    fields = config.get("fields", [])
    label_by_name = {f["name"]: f.get("label", f["name"]) for f in fields}
    headers = [label_by_name.get(col, col) for col in columns]

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    if os.path.exists(output_path):
        wb = load_workbook(output_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    # Track written rows for dedup: {sheet_name: set of (account_holder, account_number)}
    written = {}

    success = 0
    skipped = 0
    failed = 0

    for result in results:
        if result.get("error"):
            failed += 1
            logger.warning(f"Failed: {result.get('file_path')} — {result.get('error')}")
            continue

        validated_data = result.get("validated_data", {})
        statement_month = result.get("statement_month") or fallback_month or "Unsorted"

        # Get or create sheet
        if statement_month not in wb.sheetnames:
            ws = wb.create_sheet(title=statement_month)
            ws.append(headers)
            ws.freeze_panes = "A2"
            written[statement_month] = set()
        else:
            ws = wb[statement_month]
            if statement_month not in written:
                # Sheet existed before this run — load existing keys
                ah_idx = columns.index("account_holder") if "account_holder" in columns else None
                an_idx = columns.index("account_number") if "account_number" in columns else None
                written[statement_month] = set()
                for row in ws.iter_rows(min_row=2, values_only=True):
                    holder = row[ah_idx] if ah_idx is not None else None
                    number = row[an_idx] if an_idx is not None else None
                    if holder:
                        written[statement_month].add((holder, number))

        account_holder = validated_data.get("account_holder")
        account_number = validated_data.get("account_number")
        key = (account_holder, account_number)

        if key in written[statement_month]:
            skipped += 1
            logger.info(f"Skipping duplicate — {account_holder} / {account_number} in {statement_month}")
            continue

        data_row = [validated_data.get(col) for col in columns]
        ws.append(data_row)
        written[statement_month].add(key)
        success += 1

    # Auto-fit column widths
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for i, col_name in enumerate(columns, start=1):
            label = label_by_name.get(col_name, col_name)
            all_values = [label]
            for row in ws.iter_rows(min_row=2, values_only=True):
                cell_value = row[i - 1]
                if cell_value is not None:
                    all_values.append(str(cell_value))
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = int(max(len(v) for v in all_values) * 1.2) + 4

    wb.save(output_path)
    return success, skipped, failed

def main(argv=None):
    args = parse_args(argv)
    pdf_files = sorted(
        glob.glob(os.path.join(args.input, "*.pdf")) + glob.glob(os.path.join(args.input, "*.PDF"))
    )

    if not pdf_files:
        print(f"No PDFs found in {args.input}")
        return

    print(f"Found {len(pdf_files)} PDFs. Processing with {args.workers} workers...")

    results = []
    completed = 0

    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {executor.submit(process_pdf, pdf): pdf for pdf in pdf_files}
        for future in as_completed(futures):
            result = future.result()
            results.append(result)
            completed += 1
            pdf_name = os.path.basename(futures[future])
            if result.get("error"):
                print(f"[{completed}/{len(pdf_files)}] FAILED  {pdf_name}: {result.get('error')}")
            else:
                data = result.get("validated_data", {})
                print(f"[{completed}/{len(pdf_files)}] OK      {pdf_name}: {data.get('account_holder')} | {data.get('closing_balance')}")

    # No interactive prompt: an unattended run must never block on stdin.
    # Rows without a parseable date go to --month if given, else "Unsorted".
    success, skipped, failed = write_excel(results, args.month, args.output)

    print(f"\nDone. Output: {args.output}")
    print(f"  Written:   {success}")
    print(f"  Skipped:   {skipped} (duplicates)")
    print(f"  Failed:    {failed}")

if __name__ == "__main__":
    main()