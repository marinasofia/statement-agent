from openpyxl import load_workbook

from core.excel import write_workbook


def result(holder, number, month="2026-08", balance=10.0, status="OK", reasons=None, delta=None, file="f.pdf", date=None):
    return {
        "format_id": "default",
        "file_path": f"/uploads/{file}",
        "status": status,
        "review_reasons": reasons or [],
        "reconciliation": {"balance_delta": delta},
        "statement_month": month,
        "validated_data": {
            "account_holder": holder,
            "account_number": number,
            "closing_balance": balance,
            "currency": "USD",
            "bank_name": "ACME",
            "statement_date": date or (f"{month}-31" if month else None),
        },
    }


def test_rows_group_by_month_and_dedup_across_runs(tmp_path, monkeypatch):
    monkeypatch.setenv("CLIENT_ID", "default")
    out = tmp_path / "book.xlsx"

    counts = write_workbook(
        [result("Jane Doe", "111"), result("Jane Doe", "111"), result("Bob", "222", month="2026-07"),
         {"error": "boom"}],
        str(out), "default",
    )
    assert counts == {"ok": 2, "needs_review": 0, "duplicate": 1, "failed": 1}

    # Second run: the same account in the same month is a duplicate of a row
    # that already exists on disk, and a row without a date uses --month.
    counts = write_workbook(
        [result("Jane Doe", "111"), result("Zed", None, month=None)],
        str(out), "default", fallback_month="2026-09",
    )
    assert counts == {"ok": 1, "needs_review": 0, "duplicate": 1, "failed": 0}

    wb = load_workbook(out)
    assert sorted(wb.sheetnames) == ["2026-07", "2026-08", "2026-09"]
    assert wb["2026-08"].max_row == 2  # header + Jane once
    assert wb["2026-08"]["A1"].value == "Account Holder"


def test_unreconciled_rows_go_to_the_review_sheet_with_reasons(tmp_path, monkeypatch):
    monkeypatch.setenv("CLIENT_ID", "default")
    out = tmp_path / "book.xlsx"
    counts = write_workbook(
        [result("Jane Doe", "111"),
         result("Ann Off", "333", status="NEEDS_REVIEW", reasons=["balance_arithmetic: delta +12.00"], delta=12.0, file="ann.pdf")],
        str(out), "default",
    )
    assert counts == {"ok": 1, "needs_review": 1, "duplicate": 0, "failed": 0}
    wb = load_workbook(out)
    assert "Needs review" in wb.sheetnames
    ws = wb["Needs review"]
    header = [c.value for c in ws[1]]
    assert header[-4:-1] == ["Reasons", "Balance delta", "Source file"]
    row = [c.value for c in ws[2]]
    assert row[0] == "Ann Off"
    assert row[-4:-1] == ["balance_arithmetic: delta +12.00", 12.0, "ann.pdf"]
    # The reconciled row is not on the review sheet and the review row is not on the month sheet.
    assert wb["2026-08"].max_row == 2


def test_same_holder_different_account_or_balance_is_not_a_duplicate(tmp_path, monkeypatch):
    # The default client only displays holder and balance. Dedup must still
    # tell two accounts apart and two statements for one account apart.
    monkeypatch.setenv("CLIENT_ID", "default")
    out = tmp_path / "book.xlsx"
    counts = write_workbook(
        [result("Alex Morgan", "111", balance=100.0),
         result("Alex Morgan", "222", balance=100.0),            # second account
         result("Alex Morgan", "111", balance=250.0),            # different statement, same account
         result("ALEX  MORGAN", "11-1", balance=100.0)],         # same statement, formatting noise
        str(out), "default",
    )
    assert counts["ok"] == 3 and counts["duplicate"] == 1
    ws = load_workbook(out)["2026-08"]
    header = [c.value for c in ws[1]]
    assert header[-1] == "Row key"
    assert ws.column_dimensions[ws.cell(row=1, column=len(header)).column_letter].hidden
