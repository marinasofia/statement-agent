"""Regression tests for the workbook boundary, using synthetic data only."""

import math
from unittest.mock import patch

import pytest
from openpyxl import load_workbook
from pydantic import ValidationError

from agents.statement_extraction.schema import StatementData, Transaction
from core.excel import write_workbook
from tests.test_excel import result


@pytest.mark.parametrize("text", ["=1+1", "+1+1", "-1+1", "@SUM(A1)", "#N/A"])
def test_text_stays_literal_in_saved_cells(tmp_path, text):
    target = tmp_path / "book.xlsx"
    write_workbook([result(text, "111")], str(target), "default")
    workbook = load_workbook(target)
    try:
        cell = workbook["2026-08"]["A2"]
        assert cell.value == text
        assert cell.data_type == "s"
    finally:
        workbook.close()


def test_review_reasons_and_filenames_are_literal(tmp_path):
    target = tmp_path / "book.xlsx"
    row = result(
        "Example", "111", status="NEEDS_REVIEW", reasons=["=1+1"], file="=1+1.pdf"
    )
    write_workbook([row], str(target), "default")
    workbook = load_workbook(target)
    try:
        cells = list(workbook["Needs review"][2])
        assert not any(cell.data_type == "f" for cell in cells)
        assert any(cell.value == "=1+1" for cell in cells)
    finally:
        workbook.close()


@pytest.mark.parametrize("rows", [[], [{"error": "Synthetic failure"}]])
def test_empty_or_all_failed_batch_writes_valid_summary(tmp_path, rows):
    target = tmp_path / "book.xlsx"
    counts = write_workbook(rows, str(target), "default")
    workbook = load_workbook(target)
    try:
        assert workbook.sheetnames == ["Run summary"]
        assert dict(workbook.active.iter_rows(min_row=2, values_only=True)) == counts
    finally:
        workbook.close()


def test_bare_output_filename(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    write_workbook([result("Example", "111")], "book.xlsx", "default")
    assert (tmp_path / "book.xlsx").is_file()


def test_failed_save_preserves_existing_bytes_and_removes_tempfile(tmp_path):
    target = tmp_path / "book.xlsx"
    write_workbook([result("First", "111")], str(target), "default")
    original = target.read_bytes()
    with patch(
        "openpyxl.workbook.workbook.Workbook.save", side_effect=OSError("Disk failure")
    ):
        with pytest.raises(OSError):
            write_workbook([result("Second", "222")], str(target), "default")
    assert target.read_bytes() == original
    assert sorted(p.name for p in tmp_path.iterdir()) == ["book.xlsx"]


@pytest.mark.parametrize("value", [math.nan, math.inf, -math.inf, True])
def test_nonfinite_or_boolean_money_is_rejected(value):
    with pytest.raises(ValidationError):
        StatementData(account_holder="Example", closing_balance=value)
    with pytest.raises(ValidationError):
        StatementData(
            account_holder="Example", closing_balance=1, opening_balance=value
        )
    with pytest.raises(ValidationError):
        Transaction(description="Example", amount=value)


def test_incompatible_existing_headers_do_not_receive_misaligned_data(tmp_path):
    target = tmp_path / "book.xlsx"
    write_workbook([result("First", "111")], str(target), "default")
    workbook = load_workbook(target)
    workbook["2026-08"]["A1"] = "Unexpected header"
    workbook.save(target)
    workbook.close()
    original = target.read_bytes()
    with pytest.raises(ValueError, match="incompatible columns"):
        write_workbook([result("Second", "222")], str(target), "default")
    assert target.read_bytes() == original


def test_unrelated_summary_sheet_is_preserved(tmp_path):
    from openpyxl import Workbook

    target = tmp_path / "book.xlsx"
    workbook = Workbook()
    workbook.active.title = "Run summary"
    workbook.active["A1"] = "User-authored summary"
    workbook.save(target)
    workbook.close()
    write_workbook([result("Example", "111")], str(target), "default")
    workbook = load_workbook(target)
    try:
        assert workbook["Run summary"]["A1"].value == "User-authored summary"
        assert "2026-08" in workbook.sheetnames
    finally:
        workbook.close()


def test_generated_summary_is_replaced_when_statement_rows_arrive(tmp_path):
    target = tmp_path / "book.xlsx"
    write_workbook([], str(target), "default")
    write_workbook([result("Example", "111")], str(target), "default")
    workbook = load_workbook(target)
    try:
        assert workbook.sheetnames == ["2026-08"]
    finally:
        workbook.close()
